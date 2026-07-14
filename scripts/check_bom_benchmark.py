#!/usr/bin/env python3
"""
Fetch real disassembly-based BOM data as a face-validity anchor for bom_graph.py's synthetic
tier/fan-out generation.

Source: Babbitt, Madaka, Althaf, Kasulaitis & Ryen (2020), "Disassembly-based bill of materials
data for consumer electronic products," Scientific Data 7:207, doi:10.1038/s41597-020-0573-9.
Raw data: figshare 10.6084/m9.figshare.11306792 (CC0), 95 real consumer electronics products
across 25 categories, physically disassembled and measured in a lab.

This does NOT feed the generator. We verified directly against the raw "Disassembly Detail"
workbook (not just the paper's abstract) that it records mass composition per named major
assembly (e.g. Battery, Casing, Display, Motherboard) but does NOT report component counts per
assembly, hierarchy depth, or fan-out -- so it cannot calibrate BOM_DEPTH_MIN/MAX or
BOM_FANOUT_MIN/MAX directly (those describe a different thing: components consumed per assembly
in a multi-tier *supply chain* BOM, not named assemblies within one *finished product*).

What this script computes instead: the real distribution of "major assemblies per product" from
the study's own primary lab measurements (excluding its secondary literature-sourced entries) --
a genuine, citable real number, offered as face-validity context (real products decompose into
single-digit-to-low-teens major groupings) rather than a precise calibration.

Run: uv run python scripts/check_bom_benchmark.py
Writes: outputs/bom_benchmark_disassembly.md
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pandas as pd
import requests

ROOT = Path(__file__).resolve().parent.parent

FIGSHARE_ARTICLE_ID = "11306792"
DISASSEMBLY_DETAIL_FILE_ID = "22858376"
DOI = "10.1038/s41597-020-0573-9"
DATA_DOI = "10.6084/m9.figshare.11306792"


def fetch_disassembly_workbook(cache_path: Path) -> Path:
    if cache_path.exists():
        return cache_path
    url = f"https://ndownloader.figshare.com/files/{DISASSEMBLY_DETAIL_FILE_ID}"
    r = requests.get(url, timeout=60, headers={"User-Agent": "Mozilla/5.0 (compatible; research-data-fetch/1.0)"})
    r.raise_for_status()
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_bytes(r.content)
    return cache_path


def parse_assemblies_per_product(xlsx_path: Path) -> pd.DataFrame:
    """Count major-assembly rows per product, restricted to the study's own lab-disassembly
    section of each worksheet (excludes secondary literature-sourced entries in the same sheets)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    records: list[tuple[str, str, int, str]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section: str | None = None
        current_product: str | None = None
        current_count = 0
        products_in_sheet: list[tuple[str, int, str]] = []

        for row in ws.iter_rows(values_only=True):
            c0 = str(row[0]).strip() if row[0] is not None else ""
            c1 = str(row[1]).strip() if row[1] is not None else ""

            if "data from laboratory disassembly" in c0.lower():
                section = "lab"
                continue
            if "data from literature" in c0.lower():
                section = "literature"
                continue
            if c0.startswith("Product name"):
                continue

            if c1 == "" or "total mass" in c1.lower():
                if current_product is not None and current_count > 0:
                    products_in_sheet.append((current_product, current_count, section or "unknown"))
                current_product, current_count = None, 0
                continue

            if c0 != "":
                if current_product is not None and current_count > 0:
                    products_in_sheet.append((current_product, current_count, section or "unknown"))
                current_product, current_count = c0, 1
            elif current_product is not None:
                current_count += 1

        if current_product is not None and current_count > 0:
            products_in_sheet.append((current_product, current_count, section or "unknown"))

        for prod, n, sec in products_in_sheet:
            records.append((sheet_name.strip(), prod, n, sec))

    return pd.DataFrame(records, columns=["category", "product", "n_assemblies", "section"])


def main() -> None:
    cache_path = ROOT / "outputs" / "_cache" / "disassembly_detail_babbitt2020.xlsx"
    print(f"Fetching real disassembly BOM workbook (figshare {FIGSHARE_ARTICLE_ID})...")
    xlsx_path = fetch_disassembly_workbook(cache_path)

    df = parse_assemblies_per_product(xlsx_path)
    lab = df[df["section"] == "lab"]

    n = len(lab)
    mean_n = float(lab["n_assemblies"].mean())
    median_n = float(lab["n_assemblies"].median())
    lo, hi = int(lab["n_assemblies"].min()), int(lab["n_assemblies"].max())

    fanout_lo, fanout_hi = 3, 8  # bom_graph.py's BOM_FANOUT_MIN/MAX, hardcoded here for the note
    # below since this script doesn't import generate_synthetic_datasets.

    lines = [
        "# BOM structure scale check (real disassembly data)\n",
        "\n**Not a depth/fan-out calibration.** We verified directly against the raw workbook "
        "(not just the paper's abstract) that this dataset records mass composition per named "
        "major assembly but does NOT report component counts per assembly or hierarchy depth -- "
        "it cannot calibrate `BOM_DEPTH_MIN/MAX` or `BOM_FANOUT_MIN/MAX` directly (those describe "
        "components consumed per assembly across a multi-tier *supply chain*, not named assemblies "
        "within one *finished product*). This is a face-validity scale check only.\n",
        f"\n## Real data (Babbitt et al. 2020, Scientific Data, doi:{DOI}; "
        f"raw data doi:{DATA_DOI}, CC0)\n",
        f"- {n} real, lab-disassembled products (primary measurements only, excluding the study's "
        "own secondary literature-sourced entries) across 25 consumer electronics categories.\n",
        f"- Major assemblies per product: mean **{mean_n:.2f}**, median **{median_n:.0f}**, "
        f"range **{lo}-{hi}** (e.g. basic mobile phones/smartphones: 2; laptops: ~12; "
        "traditional desktops: 13).\n",
        "\n## Scale comparison\n",
        f"- `bom_graph.py`'s `BOM_FANOUT_MIN/MAX` = **{fanout_lo}-{fanout_hi}** components consumed "
        "per assembly (a different metric, but the same order of magnitude: single digits, not "
        "tens or hundreds).\n",
        "- **Conclusion**: real consumer products decompose into single-digit-to-low-teens major "
        "groupings, broadly consistent with (not a precise calibration of) the synthetic "
        "generator's single-digit fan-out range. Offered as qualitative real-data context in "
        "`data_dictionary.md`, not a validated calibration.\n",
    ]
    out_path = ROOT / "outputs" / "bom_benchmark_disassembly.md"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("".join(lines), encoding="utf-8")
    print(f"Wrote {out_path}")
    print("".join(lines))


if __name__ == "__main__":
    main()
